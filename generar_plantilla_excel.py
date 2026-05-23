"""
generar_plantilla_excel.py

Genera PLANTILLA_OFERTAS_HG.xlsx — Excel con índice, hoja por oferta,
desplegables, validaciones, protección de estructura.

Híbrido (decisión Pablo, 2026-05-21):
- Hoja Índice resumen
- Hoja Instrucciones
- Hoja Oferta_NUEVA (plantilla limpia, se duplica para cada oferta)
- Hoja Listas (oculta, datos para desplegables)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# Paleta HG
VERDE_HG = "5EBA9E"
AZUL_HG = "1B3A5C"
GRIS_FONDO = "F5F5F2"
GRIS_CLARO = "EDEEEC"
BLANCO = "FFFFFF"

# ====================================================================
# DATOS PARA LISTAS
# ====================================================================

PERFILES = ["Medicina", "Enfermería (DUE)", "TES", "TCAE", "Fisioterapeuta", "Gerocultor", "Otro"]

TIPOS_CONTRATO = [
    "Indefinido", "Temporal", "Fijo discontinuo",
    "Sustitución", "Refuerzo", "Por obra y servicio", "Prácticas", "Otro"
]

URGENCIA = ["Normal", "Alta"]
SI_NO = ["Sí", "No"]

# CCAA y provincias mapeadas
CCAA_PROVINCIAS = {
    "Andalucía": ["Almería", "Cádiz", "Córdoba", "Granada", "Huelva", "Jaén", "Málaga", "Sevilla"],
    "Aragón": ["Huesca", "Teruel", "Zaragoza"],
    "Asturias": ["Asturias"],
    "Islas Baleares": ["Illes Balears"],
    "Canarias": ["Las Palmas", "Santa Cruz de Tenerife"],
    "Cantabria": ["Cantabria"],
    "Castilla y León": ["Ávila", "Burgos", "León", "Palencia", "Salamanca", "Segovia", "Soria", "Valladolid", "Zamora"],
    "Castilla-La Mancha": ["Albacete", "Ciudad Real", "Cuenca", "Guadalajara", "Toledo"],
    "Cataluña": ["Barcelona", "Girona", "Lleida", "Tarragona"],
    "Comunidad Valenciana": ["Alicante", "Castellón", "Valencia"],
    "Extremadura": ["Badajoz", "Cáceres"],
    "Galicia": ["A Coruña", "Lugo", "Ourense", "Pontevedra"],
    "Madrid": ["Madrid"],
    "Región de Murcia": ["Murcia"],
    "Navarra": ["Navarra"],
    "País Vasco": ["Álava", "Guipúzcoa", "Vizcaya"],
    "La Rioja": ["La Rioja"],
    "Ceuta": ["Ceuta"],
    "Melilla": ["Melilla"],
}

# Lista provisional de localidades habituales (basadas en ofertas existentes)
LOCALIDADES_HABITUALES = [
    "Algeciras", "Alicante", "Almería", "Barcelona", "Benidorm", "Bilbao", "Burgos",
    "Cádiz", "Castellón", "Córdoba", "A Coruña", "Granada", "Guadalajara",
    "Huelva", "Jerez de la Frontera", "Las Palmas de Gran Canaria", "Lorca", "Madrid",
    "Málaga", "Mallorca", "Marbella", "Miranda de Ebro", "Murcia", "Pamplona",
    "Priego de Córdoba", "Santander", "San Sebastián", "Sevilla", "Tarragona",
    "Tenerife", "Tudela", "Valencia", "Valladolid", "Vigo", "Vitoria", "Zaragoza",
    "Otra (escribir manualmente)",
]

PATRONES_DIAS = [
    "L-V (lunes a viernes)",
    "L-D (todos los días)",
    "S-D (fin de semana)",
    "M-V (martes a viernes)",
    "L-S (lunes a sábado)",
    "L", "M", "X", "J", "V", "S", "D",
    "L, X, V", "M, J", "S, D",
    "Otro (escribir manualmente)",
]

# ====================================================================
# ESTILOS
# ====================================================================

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=AZUL_HG)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=AZUL_HG)
SECTION_FILL = PatternFill("solid", fgColor=GRIS_CLARO)
LABEL_FONT = Font(name="Calibri", size=10, bold=True)
INPUT_FILL = PatternFill("solid", fgColor=BLANCO)
HINT_FONT = Font(name="Calibri", size=9, italic=True, color="888888")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Protección: por defecto todo bloqueado. Las celdas de datos las desbloqueamos.
LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)

# ====================================================================
# WORKBOOK
# ====================================================================

wb = Workbook()
wb.remove(wb.active)  # eliminar hoja por defecto

# --------------------------------------------------------------------
# HOJA LISTAS (oculta) — datos para desplegables
# --------------------------------------------------------------------
ws_l = wb.create_sheet("Listas")

# Cada columna = una lista
def write_list(ws, col_letter, header, items):
    ws[f"{col_letter}1"] = header
    ws[f"{col_letter}1"].font = HEADER_FONT
    ws[f"{col_letter}1"].fill = HEADER_FILL
    for i, item in enumerate(items, start=2):
        ws[f"{col_letter}{i}"] = item

write_list(ws_l, "A", "Perfiles", PERFILES)
write_list(ws_l, "B", "Tipos contrato", TIPOS_CONTRATO)
write_list(ws_l, "C", "Urgencia", URGENCIA)
write_list(ws_l, "D", "Si No", SI_NO)
write_list(ws_l, "E", "CCAA", list(CCAA_PROVINCIAS.keys()))
write_list(ws_l, "F", "Localidades habituales", LOCALIDADES_HABITUALES)
write_list(ws_l, "G", "Patrones días", PATRONES_DIAS)

# Todas las provincias en columna H (lista plana, sin filtro por CCAA)
all_provincias = []
for prov_list in CCAA_PROVINCIAS.values():
    all_provincias.extend(prov_list)
write_list(ws_l, "H", "Provincias", sorted(set(all_provincias)))

# Provincias por CCAA en columnas separadas (para cascading futuro)
# Por simplicidad v1, usamos la lista plana de provincias. Si vemos
# que es necesario filtrar por CCAA, lo añadiremos en v1.1 con INDIRECT.

# Definir Named Ranges (necesarios para que la validación en otra hoja
# pueda referenciar las listas)
def add_range(name, ref):
    dn = DefinedName(name, attr_text=ref)
    wb.defined_names[name] = dn

add_range("Lista_Perfiles", f"Listas!$A$2:$A${len(PERFILES)+1}")
add_range("Lista_Contratos", f"Listas!$B$2:$B${len(TIPOS_CONTRATO)+1}")
add_range("Lista_Urgencia", f"Listas!$C$2:$C${len(URGENCIA)+1}")
add_range("Lista_SiNo", f"Listas!$D$2:$D${len(SI_NO)+1}")
add_range("Lista_CCAA", f"Listas!$E$2:$E${len(CCAA_PROVINCIAS)+1}")
add_range("Lista_Localidades", f"Listas!$F$2:$F${len(LOCALIDADES_HABITUALES)+1}")
add_range("Lista_PatronesDias", f"Listas!$G$2:$G${len(PATRONES_DIAS)+1}")
add_range("Lista_Provincias", f"Listas!$H$2:$H${len(set(all_provincias))+1}")

ws_l.column_dimensions["A"].width = 22
ws_l.column_dimensions["B"].width = 22
ws_l.column_dimensions["C"].width = 12
ws_l.column_dimensions["D"].width = 8
ws_l.column_dimensions["E"].width = 22
ws_l.column_dimensions["F"].width = 30
ws_l.column_dimensions["G"].width = 26
ws_l.column_dimensions["H"].width = 22

ws_l.sheet_state = "hidden"

# --------------------------------------------------------------------
# HOJA INSTRUCCIONES
# --------------------------------------------------------------------
ws_i = wb.create_sheet("Instrucciones")

instrucciones = [
    ("CÓMO USAR ESTA PLANTILLA", "title"),
    ("", ""),
    ("1.  Vaya a la pestaña 'Oferta_NUEVA' y haga clic derecho sobre la pestaña → 'Mover o copiar...'", "p"),
    ("    → marque 'Crear una copia' y acepte. Tendrá una nueva hoja 'Oferta_NUEVA (2)'.", "p"),
    ("    Renómbrela con el slug de la oferta (ej. 'DUE_Cadiz_NOCHE_2026').", "p"),
    ("", ""),
    ("2.  Rellene los campos con fondo blanco. Los demás están protegidos.", "p"),
    ("    Campos en gris claro son etiquetas; no se editan.", "p"),
    ("", ""),
    ("3.  En los campos con desplegable (Perfil, Provincia, Localidad, etc.) seleccione un valor.", "p"),
    ("    Si la opción que necesita no aparece, escriba 'Otra' o el valor manualmente.", "p"),
    ("", ""),
    ("4.  Sección HORARIO: cada fila representa un bloque de días con su horario común.", "p"),
    ("    Use varias filas si hay turnos distintos según día.", "p"),
    ("    Si no hay turno partido, rellene solo Mañana DESDE y Tarde HASTA en la misma fila.", "p"),
    ("", ""),
    ("5.  Cuando termine la hoja, vaya a 'Índice' y añada una fila resumiendo la oferta.", "p"),
    ("", ""),
    ("6.  Guarde el archivo y envíelo a Pablo (direccion@healthgroup.es).", "p"),
    ("    Pablo lo publicará en la web. Tiempo estimado: minutos.", "p"),
    ("", ""),
    ("CONTACTO", "title"),
    ("", ""),
    ("Para dudas: direccion@healthgroup.es · 952 22 45 54", "p"),
    ("", ""),
    ("PLANTILLA v1.0 · Health Group · Mayo 2026", "footer"),
]

ws_i.column_dimensions["A"].width = 120
for i, (texto, tipo) in enumerate(instrucciones, start=1):
    cell = ws_i.cell(row=i, column=1, value=texto)
    if tipo == "title":
        cell.font = Font(name="Calibri", size=14, bold=True, color=AZUL_HG)
    elif tipo == "footer":
        cell.font = HINT_FONT
    else:
        cell.font = Font(name="Calibri", size=11)
    cell.alignment = ALIGN_LEFT

# Bloquear todas las celdas (no se debe editar instrucciones)
for row in ws_i.iter_rows():
    for c in row:
        c.protection = LOCKED
ws_i.protection.sheet = True
ws_i.protection.password = "hg2026"

# --------------------------------------------------------------------
# HOJA ÍNDICE
# --------------------------------------------------------------------
ws_idx = wb.create_sheet("Índice")

ws_idx["A1"] = "ÍNDICE DE OFERTAS PUBLICADAS"
ws_idx["A1"].font = Font(name="Calibri", size=14, bold=True, color=AZUL_HG)
ws_idx.merge_cells("A1:G1")

headers_idx = [
    "Nombre de hoja", "Título oferta", "Perfil", "Provincia",
    "Localidad", "Fecha solicitud", "Estado"
]
for col, h in enumerate(headers_idx, start=1):
    c = ws_idx.cell(row=3, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER

# Anchura columnas
widths = [25, 35, 22, 18, 25, 14, 16]
for i, w in enumerate(widths):
    ws_idx.column_dimensions[get_column_letter(i+1)].width = w

# Ejemplo de fila
example = [
    "Oferta_NUEVA",
    "(Ejemplo) DUE Aplicación contrastes Miranda Ebro",
    "Enfermería (DUE)",
    "Burgos",
    "Miranda de Ebro",
    "21/05/2026",
    "Publicada",
]
for col, val in enumerate(example, start=1):
    c = ws_idx.cell(row=4, column=col, value=val)
    c.font = HINT_FONT
    c.alignment = ALIGN_LEFT
    c.border = THIN_BORDER

# Validación columna Estado
dv_estado = DataValidation(
    type="list",
    formula1='"Pendiente,Enviada a Pablo,Publicada,Archivada"',
    allow_blank=True
)
dv_estado.add(f"G4:G50")
ws_idx.add_data_validation(dv_estado)

# Desbloquear filas de datos (4-50) para que el equipo añada
for row_num in range(4, 51):
    for col_num in range(1, 8):
        ws_idx.cell(row=row_num, column=col_num).protection = UNLOCKED

ws_idx.protection.sheet = True
ws_idx.protection.password = "hg2026"
ws_idx.freeze_panes = "A4"

# --------------------------------------------------------------------
# HOJA OFERTA_NUEVA — la plantilla rellenable
# --------------------------------------------------------------------
ws_o = wb.create_sheet("Oferta_NUEVA")

# Anchuras
ws_o.column_dimensions["A"].width = 28
ws_o.column_dimensions["B"].width = 45
ws_o.column_dimensions["C"].width = 18
ws_o.column_dimensions["D"].width = 18
ws_o.column_dimensions["E"].width = 18
ws_o.column_dimensions["F"].width = 18

# Helper para escribir sección
row = 1

def section(title):
    global row
    c = ws_o.cell(row=row, column=1, value=title)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = ALIGN_LEFT
    ws_o.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

def label_input(label, hint="", dv_name=None, input_height=None, span=1):
    """Crea fila label (A) + input (B..). Devuelve la celda input principal."""
    global row
    lc = ws_o.cell(row=row, column=1, value=label)
    lc.font = LABEL_FONT
    lc.alignment = ALIGN_TOP
    lc.fill = SECTION_FILL
    lc.border = THIN_BORDER

    # Celda de input (B y siguientes según span)
    ic = ws_o.cell(row=row, column=2)
    ic.fill = INPUT_FILL
    ic.alignment = ALIGN_TOP
    ic.border = THIN_BORDER
    ic.protection = UNLOCKED

    if span > 1:
        ws_o.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1+span)
        for i in range(3, 2+span):
            ws_o.cell(row=row, column=i).border = THIN_BORDER

    # Validación
    if dv_name:
        dv = DataValidation(type="list", formula1=f"={dv_name}", allow_blank=True)
        dv.error = ""
        dv.errorTitle = ""
        dv.showErrorMessage = False  # warning style, permite valor manual
        ws_o.add_data_validation(dv)
        dv.add(ic.coordinate)

    if hint:
        # hint debajo
        hint_row = row + 1
        hc = ws_o.cell(row=hint_row, column=2, value=hint)
        hc.font = HINT_FONT
        hc.alignment = ALIGN_LEFT
        if span > 1:
            ws_o.merge_cells(start_row=hint_row, start_column=2, end_row=hint_row, end_column=1+span)
        row += 2
    else:
        row += 1

    if input_height:
        ws_o.row_dimensions[ic.row].height = input_height

    return ic

# === SECCIÓN 1: SOLICITANTE ===
section("1. DATOS DEL SOLICITANTE")
label_input("Persona que solicita:", "Tu nombre y departamento", span=5)
label_input("Fecha de solicitud:", "Formato DD/MM/AAAA", span=5)
label_input("Urgencia:", "", dv_name="Lista_Urgencia", span=5)
row += 1

# === SECCIÓN 2: DATOS PRINCIPALES ===
section("2. DATOS PRINCIPALES")
label_input("Título de la oferta:", 'Breve y descriptivo (ej. "DUE Refuerzo Verano Cádiz"). Máx ~8 palabras.', span=5)
label_input("Perfil profesional:", "Selecciona del desplegable", dv_name="Lista_Perfiles", span=5)
label_input("Empleo concreto:", 'Cargo específico (ej. "Médico de urgencias", "Auxiliar de enfermería")', span=5)
label_input("Comunidad autónoma:", "", dv_name="Lista_CCAA", span=5)
label_input("Provincia:", "Selecciona del desplegable; coherente con la CCAA", dv_name="Lista_Provincias", span=5)
label_input("Localidad concreta:", 'Si tu localidad no está en la lista, elige "Otra" y escríbela manualmente debajo', dv_name="Lista_Localidades", span=5)
label_input("Localidad (si Otra):", "Solo si arriba seleccionaste 'Otra'", span=5)
label_input("Tipo de contrato:", "", dv_name="Lista_Contratos", span=5)
label_input("Duración:", 'Ej. "3 meses", "Permanente", "Servicios puntuales 2-3 por semana"', span=5)
label_input("Nº de vacantes:", "Solo número entero", span=5)
label_input("Fecha de inicio prevista:", 'Formato DD/MM/AAAA o "Inmediata"', span=5)
label_input("Fecha de cierre candidaturas:", "(opcional) Formato DD/MM/AAAA", span=5)
label_input("Salario:", 'Ej. "20-22 €/h brutos", "1.800 €/mes brutos", "Según convenio"', span=5)
row += 1

# === SECCIÓN 3: HORARIO ===
section("3. HORARIO SEMANAL")

# Headers de la tabla
header_row = row
ws_o.cell(row=row, column=1, value="Días").font = LABEL_FONT
ws_o.cell(row=row, column=1).fill = SECTION_FILL
ws_o.cell(row=row, column=1).border = THIN_BORDER
ws_o.cell(row=row, column=1).alignment = ALIGN_CENTER
ws_o.cell(row=row, column=2, value="Mañana DESDE").font = LABEL_FONT
ws_o.cell(row=row, column=2).fill = SECTION_FILL
ws_o.cell(row=row, column=2).border = THIN_BORDER
ws_o.cell(row=row, column=2).alignment = ALIGN_CENTER
ws_o.cell(row=row, column=3, value="Mañana HASTA").font = LABEL_FONT
ws_o.cell(row=row, column=3).fill = SECTION_FILL
ws_o.cell(row=row, column=3).border = THIN_BORDER
ws_o.cell(row=row, column=3).alignment = ALIGN_CENTER
ws_o.cell(row=row, column=4, value="Tarde DESDE").font = LABEL_FONT
ws_o.cell(row=row, column=4).fill = SECTION_FILL
ws_o.cell(row=row, column=4).border = THIN_BORDER
ws_o.cell(row=row, column=4).alignment = ALIGN_CENTER
ws_o.cell(row=row, column=5, value="Tarde HASTA").font = LABEL_FONT
ws_o.cell(row=row, column=5).fill = SECTION_FILL
ws_o.cell(row=row, column=5).border = THIN_BORDER
ws_o.cell(row=row, column=5).alignment = ALIGN_CENTER
row += 1

# Hint
ws_o.cell(row=row, column=1, value=(
    "Una fila por cada bloque de horario distinto. "
    "Si el horario es continuo (no partido), rellene solo Mañana DESDE y Tarde HASTA."
)).font = HINT_FONT
ws_o.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
ws_o.cell(row=row, column=1).alignment = ALIGN_LEFT
row += 1

# 5 filas de horario rellenables
dv_dias = DataValidation(type="list", formula1="=Lista_PatronesDias", allow_blank=True)
dv_dias.showErrorMessage = False
ws_o.add_data_validation(dv_dias)

for i in range(5):
    for col in range(1, 6):
        c = ws_o.cell(row=row, column=col)
        c.fill = INPUT_FILL
        c.border = THIN_BORDER
        c.protection = UNLOCKED
        if col >= 2:
            c.number_format = "HH:MM"
            c.alignment = ALIGN_CENTER
    dv_dias.add(ws_o.cell(row=row, column=1).coordinate)
    row += 1
row += 1

# === SECCIÓN 4: DESCRIPCIÓN ===
section("4. DESCRIPCIÓN DE LA OFERTA")

descripcion_campos = [
    ("Sobre el puesto:", 'Breve presentación del centro, el servicio o la actividad. 2-4 frases.', 100),
    ("Requisitos mínimos:", 'Obligatorios. Lista en bullets o frases cortas, uno por línea.', 120),
    ("Requisitos deseados:", 'No obligatorios pero valorables. Opcional.', 80),
    ("Lo que ofrecemos:", 'Beneficios y condiciones atractivas.', 100),
    ("Información adicional:", 'Cualquier otro detalle relevante. Opcional.', 60),
]

for label, hint, height in descripcion_campos:
    lc = ws_o.cell(row=row, column=1, value=label)
    lc.font = LABEL_FONT
    lc.fill = SECTION_FILL
    lc.border = THIN_BORDER
    lc.alignment = ALIGN_TOP

    ic = ws_o.cell(row=row, column=2)
    ic.fill = INPUT_FILL
    ic.alignment = ALIGN_TOP
    ic.border = THIN_BORDER
    ic.protection = UNLOCKED
    ws_o.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    for c2 in range(3, 7):
        ws_o.cell(row=row, column=c2).border = THIN_BORDER
    ws_o.row_dimensions[row].height = height
    row += 1

    # Hint
    hc = ws_o.cell(row=row, column=2, value=hint)
    hc.font = HINT_FONT
    hc.alignment = ALIGN_LEFT
    ws_o.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1

row += 1

# === SECCIÓN 5: IMAGEN Y OBSERVACIONES ===
section("5. IMAGEN Y OBSERVACIONES")
label_input(
    "¿Usar imagen automática?",
    'Sí = generamos imagen automática (foto perfil + mapa CCAA). No = especifica imagen abajo.',
    dv_name="Lista_SiNo", span=5
)
label_input("Imagen específica (URL):", "Solo si elegiste 'No' arriba", span=5)
ic_obs = label_input(
    "Observaciones para Pablo:",
    "Cualquier matiz, contexto interno o instrucción especial. Opcional.",
    span=5
)
ws_o.row_dimensions[ic_obs.row].height = 60

# Fila final con marca de plantilla
row += 2
foot = ws_o.cell(row=row, column=1, value="PLANTILLA OFERTA v1.0 · Health Group · 2026")
foot.font = HINT_FONT
ws_o.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

# Proteger la hoja
ws_o.protection.sheet = True
ws_o.protection.password = "hg2026"
ws_o.freeze_panes = "A2"

# --------------------------------------------------------------------
# Reordenar hojas: Índice, Instrucciones, Oferta_NUEVA, Listas (oculta)
# --------------------------------------------------------------------
wb._sheets = [
    wb["Índice"],
    wb["Instrucciones"],
    wb["Oferta_NUEVA"],
    wb["Listas"],
]

# Guardar
import os
output = os.path.join(os.path.dirname(os.path.abspath(__file__)), "PLANTILLA_OFERTAS_HG.xlsx")
wb.save(output)
print(f"OK generada: {output}")
print(f"Tamaño: {os.path.getsize(output)} bytes")
